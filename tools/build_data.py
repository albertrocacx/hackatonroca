"""
Ingesta: convierte las fuentes (Excel de productos + CSV de relaciones) en los
JSON que consume el backend. Pensado para crecer: añade aqui nuevas fuentes.

Salida -> backend/data/products.json  y  backend/data/relations.json

Uso:
  python build_data.py
"""
import openpyxl, csv, json, os, re
from collections import defaultdict

# --- Fuentes (ajusta las rutas si mueves los ficheros) ---
XLSX = r"C:\Users\malalb01\Downloads\filesPy\ROCA_productos_definitivo.xlsx"
RELS = r"C:\Users\malalb01\OneDrive - Roca Group\Escritorio\IA\relations.csv"

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(HERE, "..", "backend", "data")
os.makedirs(OUT_DIR, exist_ok=True)

CODE_PREFIX = re.compile(r"^\d{2}\.\d{2}\.\d{2}\s*-\s*")  # "00.01.00 - Lavabo" -> "Lavabo"

def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def title_of(row):
    for key in ("Description Title es_ES", "Short Description Title es_ES",
                "Marketing Description es_ES"):
        t = clean(row.get(key))
        if t:
            return t
    pd = clean(row.get("ProductDescription es_ES"))
    if pd:
        return CODE_PREFIX.sub("", pd)
    return clean(row.get("SKU"))

def raw_val(v):
    """Valor serializable a JSON tal cual viene del Excel (str/num/bool/None)."""
    if isinstance(v, (str, int, float, bool)) or v is None:
        return v
    return str(v)

# ---------- PRODUCTOS ----------
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb.active
headers = [c.value for c in ws[1]]
products = []
for r in ws.iter_rows(min_row=2, values_only=True):
    row = dict(zip(headers, r))
    # 1) TODAS las columnas del Excel, con su nombre original
    rec = {h: raw_val(row.get(h)) for h in headers}
    # 2) campos derivados/normalizados que usa el backend (no romper la app)
    rec.update({
        "sku": clean(row.get("SKU")),
        "model": clean(row.get("Model")),
        "title": title_of(row),
        "category": clean(row.get("SearchHierarchyDescription es_ES 1")),
        "subcategory": clean(row.get("SearchHierarchyDescription es_ES 2")),
        "category_base": clean(row.get("category_base")),
        "collection": clean(row.get("DesignLineName es_ES")),
        "finish": clean(row.get("Finish Description es_ES")),
        "finish_code": clean(row.get("Finish Code")),
        "price_rrp": row.get("RRP_eur"),
        "dims": {
            "length_mm": row.get("Length_mm"),
            "width_mm": row.get("Width_mm"),
            "height_mm": row.get("Height_mm"),
            "weight_kg": row.get("GrossWeight_kg"),
        },
        "is_spare_part": bool(row.get("is_spare_part")),
        "ecommerce": str(row.get("ecommerce")).lower() == "true",
        "status": clean(row.get("Status")),
        "product_code": clean(row.get("ProductCode")),
        "desc": {
            "marketing": clean(row.get("Marketing Description es_ES")),
            "extended": clean(row.get("Extended Description es_ES")),
        },
    })
    products.append(rec)
wb.close()

# ---------- RELACIONES (agrupadas por modelo) ----------
rels_loaded = os.path.exists(RELS)
relations = defaultdict(list)
if rels_loaded:
    with open(RELS, encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            relations[row["model"]].append({
                "type": row["relation"],                 # compatible | optional | included | sparepart
                "code": row["related_code"],
                "description": row.get("related_description"),
                "collection": row.get("related_designline"),
                "category": row.get("related_category"),
            })

# ---------- Guardar ----------
with open(os.path.join(OUT_DIR, "products.json"), "w", encoding="utf-8") as f:
    json.dump(products, f, ensure_ascii=False)
print(f"products.json : {len(products)} productos, {len(products[0])} campos por producto")

# relations.json solo se reescribe si se pudo leer el CSV (no vaciar lo existente)
if rels_loaded:
    with open(os.path.join(OUT_DIR, "relations.json"), "w", encoding="utf-8") as f:
        json.dump(relations, f, ensure_ascii=False)
    print(f"relations.json: {sum(len(v) for v in relations.values())} relaciones en {len(relations)} modelos")
else:
    print(f"AVISO: no encuentro {RELS} -> NO se toca relations.json (se conserva el existente).")

print(f"Salida -> {os.path.abspath(OUT_DIR)}")
